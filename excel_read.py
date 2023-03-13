# парсинг таблицы
from datetime import datetime
import pymysql
from openpyxl import load_workbook


connection = pymysql.connect(
        host='localhost',
        port=3306,
        user='root',
        password='Egor76787673',
        database='marykay_db',
        cursorclass=pymysql.cursors.DictCursor
    )


# Функция, вставляющая полученные из строки значения (pid, name, phone, date, app)
def insert(a):
    if a[1] and a[2] and a[3] and a[0] is not None:
        with connection.cursor() as cursor:
            add_info = "INSERT INTO client_base (pid, name, phone, date, app) VALUES (1, '" + a[0] + "','" + a[2] + "', '" + a[1].strftime('%Y-%m-%d') + "', '" + a[3] + "')"
            cursor.execute(add_info)
            connection.commit()


def phone_number(x):
    s = str(x)
    phone = ''
    for i in s:
        if i.isdigit():
            phone += i
    phone = phone[-10:]
    return phone


wb = load_workbook('АВАТАР клиента.xlsx')
source = wb.active
rows = []
for row in source.iter_rows(max_col=4):
    a = []
    k = -1
    for cell in row:
        k += 1
        if k == 2:
            a.append(phone_number(cell.value))
        else:
            a.append(cell.value)
    insert(a)